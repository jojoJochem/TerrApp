import io
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
TESTS_DIR = Path(__file__).resolve().parent
FIXTURES_DIR = TESTS_DIR / "fixtures"
MNT_DATA_DIR = Path("/mnt/data")

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


REAL_EXCEL_FILES = [
    "24-08-08_bijlage_2.1_1_grond_T2413196.xlsm",
    "25-06-03_bijlage_2_1_grond_T2513603.xlsm",
    "25-06-03_bijlage_2_2_grond_T2513603.xlsm",
    "2.1 A02 A02 A04.xlsm",
    "2.1 MMA01 MMA02 MMA03.xlsm",
    "2.1 D04 D05 D06.xlsm",
    "2.1 MMA04 MMA05 MMA06.xlsm",
    "2.2 A04 B09 D02 D06.xlsm",
]

REAL_GROUND_FILES = [
    "24-08-08_bijlage_2.1_1_grond_T2413196.xlsm",
    "25-06-03_bijlage_2_1_grond_T2513603.xlsm",
    "25-06-03_bijlage_2_2_grond_T2513603.xlsm",
    "2.1 A02 A02 A04.xlsm",
    "2.1 MMA01 MMA02 MMA03.xlsm",
    "2.1 D04 D05 D06.xlsm",
    "2.1 MMA04 MMA05 MMA06.xlsm",
]

REAL_GROUNDWATER_FILES = [
    "2.2 A04 B09 D02 D06.xlsm",
]


def _resolve_fixture_file(filename: str) -> Path:
    """
    Zoek testbestand eerst in tests/fixtures, daarna in /mnt/data.
    Lokaal kun je dus gewoon alle .xlsm bestanden naar tests/fixtures kopiëren.
    """
    candidates = [
        FIXTURES_DIR / filename,
        MNT_DATA_DIR / filename,
    ]

    for path in candidates:
        if path.exists():
            return path

    raise FileNotFoundError(
        f"Test fixture niet gevonden: {filename}. "
        f"Zet dit bestand in {FIXTURES_DIR} of zorg dat het in /mnt/data staat."
    )


@pytest.fixture()
def fixture_file():
    return _resolve_fixture_file


@pytest.fixture()
def real_excel_files(fixture_file):
    return [fixture_file(name) for name in REAL_EXCEL_FILES]


@pytest.fixture()
def real_ground_files(fixture_file):
    return [fixture_file(name) for name in REAL_GROUND_FILES]


@pytest.fixture()
def real_groundwater_files(fixture_file):
    return [fixture_file(name) for name in REAL_GROUNDWATER_FILES]


@pytest.fixture()
def copied_real_excel_files(tmp_path: Path, real_excel_files):
    """
    Kopieer echte fixtures naar tmp_path, zodat Flask upload tests echte file handles
    kunnen gebruiken zonder de bronbestanden te muteren.
    """
    copied = []

    for src in real_excel_files:
        dst = tmp_path / src.name
        shutil.copy2(src, dst)
        copied.append(dst)

    return copied


@pytest.fixture()
def tmp_xlsm_path(tmp_path: Path):
    return tmp_path / "input.xlsm"


def _write_cell(ws, r, c, v):
    ws.cell(row=r, column=c, value=v)


@pytest.fixture()
def make_tabel_workbook():
    """
    Bouwt een minimale grond-Excel die de parser verwacht:
    sheet 'Tabel' met anchor-rows + sample kolommen.
    """
    def _maker(
        *,
        project_code="T.25.16109",
        include_pfas=True,
        samples=("MM01", "02"),
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabel"

        _write_cell(ws, 1, 1, f"Project: {project_code}")

        row_mm = 5
        row_ms = 12
        row_rbk = 25
        row_pfas = 26
        row_tot = 27

        _write_cell(ws, row_mm, 1, "Mengmonster / boring")
        _write_cell(ws, row_ms, 1, "Monstersamenstelling")
        _write_cell(ws, row_rbk, 1, "Kwaliteitsklasse Rbk")
        _write_cell(ws, row_pfas, 1, "Kwaliteitsklasse PFAS")
        _write_cell(ws, row_tot, 1, "Kwaliteitsklasse totaal")

        start_col = 2

        for idx, header in enumerate(samples):
            col = start_col + idx
            _write_cell(ws, row_mm, col, header)

            _write_cell(ws, row_mm + 1, col, "Zand")
            _write_cell(ws, row_mm + 2, col, "Klei")
            _write_cell(ws, row_mm + 3, col, "Grind")

            _write_cell(ws, row_ms + 1, col, "01 (0,50-1,00)")
            _write_cell(ws, row_ms + 2, col, "02 (0,50-1,00)")
            _write_cell(ws, row_ms + 3, col, "03 (0,50-1,00)")
            _write_cell(ws, row_ms + 4, col, "05 (0,50-1,00)")
            _write_cell(ws, row_ms + 5, col, "zz")

            _write_cell(ws, row_rbk, col, "wonen")
            _write_cell(ws, row_pfas, col, "landbouw / natuur" if include_pfas else "")
            _write_cell(ws, row_tot, col, "klasse wonen")

        if include_pfas:
            _write_cell(ws, 40, 1, "PFOS")
            _write_cell(ws, 40, start_col, 0.12)

        _write_cell(ws, 41, 1, "Arseen")
        _write_cell(ws, 41, start_col, 3.14)

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    return _maker


@pytest.fixture()
def make_groundwater_workbook():
    """
    Bouwt een minimale grondwater-Excel voor Tabel 3.
    """
    def _maker(project_code="T.25.16109") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabel"

        _write_cell(ws, 1, 1, f"Project: {project_code}")

        row_pb = 5
        row_filter = 8
        row_stand = 9
        row_ph = 12
        row_egv = 13
        row_troebelheid = 15

        _write_cell(ws, row_pb, 1, "Peilbuis")
        _write_cell(ws, row_filter, 1, "Filterstelling (m - mv.)")
        _write_cell(ws, row_stand, 1, "Grondwaterstand (m - mv.)")
        _write_cell(ws, row_ph, 1, "pH (-)")
        _write_cell(ws, row_egv, 1, "Geleidbaarheid (µS/cm)")
        _write_cell(ws, row_troebelheid, 1, "Troebelheid (NTU)")

        samples = [
            ("D02", "2,50-3,50", 1.7, 7.6, 1400, 8.7),
            ("D06", "2,50-3,50", 1.4, 7.7, 1100, 74),
        ]

        start_col = 2

        for idx, values in enumerate(samples):
            col = start_col + idx
            peilbuis, filterstelling, stand, ph, egv, troebelheid = values

            _write_cell(ws, row_pb, col, peilbuis)
            _write_cell(ws, row_filter, col, filterstelling)
            _write_cell(ws, row_stand, col, stand)
            _write_cell(ws, row_ph, col, ph)
            _write_cell(ws, row_egv, col, egv)
            _write_cell(ws, row_troebelheid, col, troebelheid)

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    return _maker